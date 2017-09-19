# -*- coding: utf-8 -*-
import sys
import re
from datetime import datetime, date, time
import scrapy
from scrapy.spiders import CrawlSpider, Rule, Spider
from scrapy.linkextractors import LinkExtractor
from openpyxl import Workbook
from openpyxl import load_workbook
import wikipedia as w


class WikipediaSpider(CrawlSpider):
    name = "wikicrawler"
    allowed_domains = ["wikipedia.org"]
    start_urls = ['http://wikipedia.org/']

    wb = None
    wb_title = ""
    ws_out = None
    avoids = (
        "https://en.wikipedia.org/w/index.php",
        "https://en.m.wikipedia.org/",
        "https://en.wikipedia.org/w/api.php",
    )

    rules = (
        # This rule ensure that only english desktop articles will be considered, 
        # avoiding foreign languages and mobile duplicates
        Rule(LinkExtractor(allow="https://en.wikipedia.org/", deny=avoids), callback='countwords', follow=True),
    )

    def __init__(self, title='', workbook='', domain=None, *args, **kwargs):
        super(WikipediaSpider, self).__init__(*args, **kwargs)
        if title:
            self.start_urls = ['https://www.wikipedia.org/wiki/%s' % title]
        else:
            print("Please enter starting Wiki page. For example, dogs.")

        if workbook:
            self.wb_title = workbook
            try:
                self.wb = load_workbook(workbook)
            except IOError:
                print("Workbook not found. Creating new workbook with name " + self.wb_title)
                self.wb = Workbook()
            # Create new Excel worksheet with today's date
            d = datetime.now()
            time = d.strftime("%B %d %Y %I.%M%p")
            self.ws_out = self.wb.create_sheet(title=time)
            self.ws_out.append(['Names', 'Word Count', 'Links', 'Categories'])
        else:
            print("Please enter workbook name for data output.")



    def parse_item(self, response):
        item = scrapy.Item()
        return item



    def countwords(self, response):
        data = response.body
        words = len(data.split())
        
        # Get title of page
        name = response.xpath('//h1[@id="firstHeading"]/text()').extract()
        try:
            name = name[0]

        except IndexError:
            name = ""

        # Get categories that this page is under
        categories = []
        for tag in response.xpath('//div[@id="catlinks"]//li/a/text()').extract():
            categories.append(tag)

        
        # Print row for each category mapping to the page
        for category in categories:
            ar = [name, words, response.url, category]
            self.ws_out.append(ar)
            self.wb.save(self.wb_title)

        if len(categories) < 1:
            ar = [name, words, response.url, ""]
            self.ws_out.append(ar)
            self.wb.save(self.wb_title)




